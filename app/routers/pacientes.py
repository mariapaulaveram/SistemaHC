from datetime import datetime
import logging

import csv
import io
from datetime import date

from fastapi import APIRouter, Depends, Form, Query, Request, HTTPException
from fastapi.responses import RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates  # kept for type hints
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session
from sqlalchemy import or_, asc, desc
from .. import models, schemas, database
from ..auth import require_login
from ..utils import render_template, make_templates, set_flash_message, check_csrf
import os

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/pacientes", tags=["pacientes"], dependencies=[Depends(require_login)])
templates = make_templates(directory=os.path.join(os.path.dirname(__file__), "..", "templates"))


PER_PAGE = 20

@router.get("/")
@router.get("")
def pagina_pacientes(
    request: Request,
    q: str | None = Query(None, title="Buscar"),
    sort: str = Query("nombre", pattern="^(nombre|dni|fecha_nacimiento)$"),
    order: str = Query("asc", pattern="^(asc|desc)$"),
    page: int = Query(1, ge=1),
    db: Session = Depends(database.get_db),
):
    query = db.query(models.Paciente).distinct()

    if q:
        q = q.strip()
        logger.info(f"BUSQUEDA: q={repr(q)}")

        paciente_exacto = db.query(models.Paciente).filter(models.Paciente.dni == q).first()
        if paciente_exacto:
            return RedirectResponse(url=f"/pacientes/{paciente_exacto.id}", status_code=303)

        q_pattern = f"%{q}%"
        query = (
            query
            .outerjoin(models.Consulta, models.Consulta.paciente_id == models.Paciente.id)
            .filter(
                or_(
                    models.Paciente.nombre.ilike(q_pattern),
                    models.Paciente.dni.ilike(q_pattern),
                    models.Consulta.diagnostico.ilike(q_pattern),
                    models.Consulta.tipo_lesion.ilike(q_pattern),
                )
            )
        )

    order_col = getattr(models.Paciente, sort)
    order_col = desc(order_col) if order == "desc" else asc(order_col)
    query = query.order_by(order_col)

    total = query.count()
    import math
    total_pages = max(1, math.ceil(total / PER_PAGE))
    page = min(page, total_pages)
    pacientes = query.offset((page - 1) * PER_PAGE).limit(PER_PAGE).all()

    return render_template(
        templates, request, "lista_pacientes.html",
        {
            "pacientes": pacientes,
            "search_params": {"q": q, "sort": sort, "order": order},
            "page": page,
            "total_pages": total_pages,
            "total": total,
            "per_page": PER_PAGE,
        },
    )


@router.get("/lista")
def listar_pacientes_web(
    request: Request,
    db: Session = Depends(database.get_db),
    limit: int | None = Query(None, ge=1, le=100),
    recent: bool = Query(False),
):
    query = db.query(models.Paciente)
    if recent:
        query = query.order_by(models.Paciente.id.desc())
    else:
        query = query.order_by(models.Paciente.nombre)
    if limit:
        query = query.limit(limit)
    pacientes = query.all()
    return templates.TemplateResponse(
        request=request,
        name="lista_pacientes_items.html",
        context={"pacientes": pacientes},
    )


def _edad(fecha_nac):
    if not fecha_nac:
        return ""
    hoy = date.today()
    return hoy.year - fecha_nac.year - ((hoy.month, hoy.day) < (fecha_nac.month, fecha_nac.day))

def _pacientes_ordenados(db):
    return db.query(models.Paciente).order_by(models.Paciente.nombre).all()

COLUMNAS = ["N°", "Nombre", "DNI", "Fecha nacimiento", "Edad", "Sexo",
            "Ocupación", "Teléfono", "Tipo de piel", "Alergias",
            "Medicaciones actuales", "Antecedentes"]

def _fila(p):
    return [
        p.id, p.nombre, p.dni or "",
        p.fecha_nacimiento.strftime("%d/%m/%Y") if p.fecha_nacimiento else "",
        _edad(p.fecha_nacimiento),
        p.sexo or "", p.ocupacion or "", p.telefono or "",
        p.tipo_piel or "", p.alergias or "",
        p.medicaciones_actuales or "", p.antecedentes or "",
    ]


@router.get("/exportar/csv")
def exportar_csv(db: Session = Depends(database.get_db)):
    pacientes = _pacientes_ordenados(db)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(COLUMNAS)
    for p in pacientes:
        writer.writerow(_fila(p))
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=pacientes.csv"},
    )


@router.get("/exportar/excel")
def exportar_excel(db: Session = Depends(database.get_db)):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    pacientes = _pacientes_ordenados(db)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pacientes"

    # Encabezado
    header_fill = PatternFill("solid", fgColor="0077A8")
    header_font = Font(bold=True, color="FFFFFF")
    for col, titulo in enumerate(COLUMNAS, 1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Datos
    for row_idx, p in enumerate(pacientes, 2):
        for col_idx, val in enumerate(_fila(p), 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
        if row_idx % 2 == 0:
            for col_idx in range(1, len(COLUMNAS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill("solid", fgColor="F0F4F8")

    # Ancho de columnas
    anchos = [6, 30, 14, 16, 6, 10, 20, 14, 12, 25, 25, 30]
    for col, ancho in enumerate(anchos, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ancho

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        iter([output.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=pacientes.xlsx"},
    )


@router.get("/nuevo")
def pagina_crear_paciente(request: Request):
    return render_template(templates, request, "crear_paciente.html", {"form_data": {}})


@router.get("/{paciente_id}/editar")
def editar_paciente_form(request: Request, paciente_id: int, db: Session = Depends(database.get_db)):
    paciente = db.query(models.Paciente).filter(models.Paciente.id == paciente_id).first()
    if not paciente:
        raise HTTPException(status_code=404, detail="Paciente no encontrado")
    return render_template(templates, request, "paciente_editar.html", {"paciente": paciente, "error": None})


@router.post("/{paciente_id}/editar")
def editar_paciente(
    request: Request,
    paciente_id: int,
    csrf_token: str = Form(default=""),
    nombre: str = Form(...),
    dni: str = Form(...),
    fecha_nacimiento: str = Form(None),
    sexo: str = Form(None),
    ocupacion: str = Form(None),
    telefono: str = Form(None),
    tipo_piel: str = Form(None),
    alergias: str = Form(None),
    medicaciones_actuales: str = Form(None),
    antecedentes: str = Form(None),
    db: Session = Depends(database.get_db),
):
    check_csrf(csrf_token, request)
    paciente = db.query(models.Paciente).filter(models.Paciente.id == paciente_id).first()
    if not paciente:
        raise HTTPException(status_code=404, detail="Paciente no encontrado")

    paciente.nombre = nombre
    paciente.dni = dni
    paciente.fecha_nacimiento = datetime.strptime(fecha_nacimiento, "%Y-%m-%d").date() if fecha_nacimiento else None
    paciente.sexo = sexo
    paciente.ocupacion = ocupacion
    paciente.telefono = telefono
    paciente.tipo_piel = tipo_piel
    paciente.alergias = alergias
    paciente.medicaciones_actuales = medicaciones_actuales
    paciente.antecedentes = antecedentes

    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        return render_template(templates, request, "paciente_editar.html", {
            "paciente": paciente,
            "error": f"Ya existe un paciente con el DNI {dni}.",
        })

    response = RedirectResponse(url=f"/pacientes/{paciente_id}", status_code=303)
    set_flash_message(response, "Paciente actualizado correctamente.")
    return response


@router.get("/{paciente_id}")
def ver_paciente(request: Request, paciente_id: int, db: Session = Depends(database.get_db)):
    paciente = db.query(models.Paciente).filter(models.Paciente.id == paciente_id).first()
    if not paciente:
        raise HTTPException(status_code=404, detail="Paciente no encontrado")
    consultas = (
        db.query(models.Consulta)
        .filter(models.Consulta.paciente_id == paciente_id)
        .order_by(models.Consulta.fecha.desc())
        .all()
    )
    return render_template(templates, request, "paciente_detalle.html", {"paciente": paciente, "consultas": consultas})


@router.post("/")
def crear_paciente_web(
    request: Request,
    csrf_token: str = Form(default=""),
    nombre: str = Form(...),
    dni: str = Form(...),
    fecha_nacimiento: str = Form(None),
    sexo: str = Form(None),
    ocupacion: str = Form(None),
    telefono: str = Form(None),
    tipo_piel: str = Form(None),
    alergias: str = Form(None),
    medicaciones_actuales: str = Form(None),
    antecedentes: str = Form(None),
    db: Session = Depends(database.get_db),
):
    check_csrf(csrf_token, request)
    nuevo = models.Paciente(
        nombre=nombre,
        dni=dni,
        fecha_nacimiento=datetime.strptime(fecha_nacimiento, "%Y-%m-%d").date() if fecha_nacimiento else None,
        sexo=sexo,
        ocupacion=ocupacion,
        telefono=telefono,
        tipo_piel=tipo_piel,
        alergias=alergias,
        medicaciones_actuales=medicaciones_actuales,
        antecedentes=antecedentes,
    )
    db.add(nuevo)
    try:
        db.commit()
        db.refresh(nuevo)
        response = RedirectResponse(url=f"/pacientes/{nuevo.id}", status_code=303)
        set_flash_message(response, "Paciente creado correctamente.")
        return response
    except IntegrityError:
        db.rollback()
        return render_template(templates, request, "crear_paciente.html", {
            "error": f"Ya existe un paciente con el DNI {dni}.",
            "form_data": {
                "nombre": nombre, "dni": dni, "fecha_nacimiento": fecha_nacimiento,
                "sexo": sexo, "ocupacion": ocupacion, "telefono": telefono,
                "tipo_piel": tipo_piel, "alergias": alergias,
                "medicaciones_actuales": medicaciones_actuales, "antecedentes": antecedentes,
            },
        })


@router.get("/{paciente_id}/imprimir")
def imprimir_paciente(request: Request, paciente_id: int, db: Session = Depends(database.get_db)):
    paciente = db.query(models.Paciente).filter(models.Paciente.id == paciente_id).first()
    if not paciente:
        raise HTTPException(status_code=404, detail="Paciente no encontrado")
    consultas = (
        db.query(models.Consulta)
        .filter(models.Consulta.paciente_id == paciente_id)
        .order_by(models.Consulta.fecha.desc())
        .all()
    )
    from datetime import date
    return render_template(templates, request, "paciente_imprimir.html", {
        "paciente": paciente,
        "consultas": consultas,
        "fecha_hoy": date.today().strftime("%d/%m/%Y"),
    })


@router.post("/{paciente_id}/eliminar")
def eliminar_paciente(
    request: Request,
    paciente_id: int,
    csrf_token: str = Form(default=""),
    db: Session = Depends(database.get_db),
):
    check_csrf(csrf_token, request)
    paciente = db.query(models.Paciente).filter(models.Paciente.id == paciente_id).first()
    if not paciente:
        raise HTTPException(status_code=404, detail="Paciente no encontrado")
    nombre = paciente.nombre
    db.delete(paciente)
    db.commit()
    response = RedirectResponse(url="/pacientes", status_code=303)
    set_flash_message(response, f"Paciente {nombre} eliminado.")
    return response


@router.post("/api", response_model=schemas.Paciente)
def crear_paciente_api(paciente: schemas.PacienteCreate, db: Session = Depends(database.get_db)):
    nuevo = models.Paciente(**paciente.model_dump())
    db.add(nuevo)
    db.commit()
    db.refresh(nuevo)
    return nuevo
